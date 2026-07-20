"""Integration tests for data and document tools.

Tests data inspection, document extraction, and tabular data analysis.
"""

import tempfile
from pathlib import Path

import pytest

# ---------------------------------------------------------------------------
# Data Inspection Tools Tests
# ---------------------------------------------------------------------------


class TestDataInspectionTools:
    """Integration tests for data file inspection tools."""

    @pytest.fixture
    def inspect_tool(self):
        """Create InspectDataTool instance."""
        from soothe_nano.toolkits.data import InspectDataTool

        return InspectDataTool()

    def test_inspect_csv_file(self, inspect_tool) -> None:
        """Test inspecting CSV file structure."""
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_file = Path(tmpdir) / "data.csv"
            csv_file.write_text("name,age,city\nAlice,30,NYC\nBob,25,LA\n")

            result = inspect_tool._run(str(csv_file))

            # Should return structure info
            assert isinstance(result, (str, dict))
            if isinstance(result, dict):
                assert "columns" in result or "shape" in result or "error" in result

    def test_inspect_json_file(self, inspect_tool) -> None:
        """Test inspecting JSON file structure."""
        with tempfile.TemporaryDirectory() as tmpdir:
            import json

            json_file = Path(tmpdir) / "data.json"
            json_file.write_text(json.dumps({"users": [{"name": "Alice"}, {"name": "Bob"}]}))

            result = inspect_tool._run(str(json_file))

            assert isinstance(result, (str, dict))

    def test_inspect_parquet_file(self, inspect_tool) -> None:
        """Test inspecting Parquet file structure."""
        pytest.importorskip("pyarrow")
        import pyarrow as pa
        import pyarrow.parquet as pq

        with tempfile.TemporaryDirectory() as tmpdir:
            # Create parquet file
            table = pa.table({"col1": [1, 2, 3], "col2": ["a", "b", "c"]})
            pq_file = Path(tmpdir) / "data.parquet"
            pq.write_table(table, str(pq_file))

            result = inspect_tool._run(str(pq_file))

            assert isinstance(result, (str, dict))

    def test_inspect_excel_file(self, inspect_tool) -> None:
        """Test inspecting Excel file structure."""
        pytest.importorskip("openpyxl")

        with tempfile.TemporaryDirectory() as tmpdir:
            import openpyxl

            xlsx_file = Path(tmpdir) / "data.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "Header"
            ws["A2"] = "Data"
            wb.save(xlsx_file)

            result = inspect_tool._run(str(xlsx_file))

            assert isinstance(result, (str, dict))


class TestDataSummaryTools:
    """Integration tests for data summarization tools."""

    @pytest.fixture
    def summarize_tool(self):
        """Create SummarizeDataTool instance."""
        from soothe_nano.toolkits.data import SummarizeDataTool

        return SummarizeDataTool()

    def test_summarize_numeric_data(self, summarize_tool) -> None:
        """Test summarizing numeric data with statistics."""
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_file = Path(tmpdir) / "numbers.csv"
            csv_file.write_text("value\n10\n20\n30\n40\n50\n")

            result = summarize_tool._run(str(csv_file))

            # Should return statistics (mean, median, std, etc.)
            assert isinstance(result, (str, dict))
            if isinstance(result, str):
                # Should contain statistical info
                assert any(
                    term in result.lower() for term in ["mean", "average", "count", "min", "max"]
                )

    def test_summarize_categorical_data(self, summarize_tool) -> None:
        """Test summarizing categorical data."""
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_file = Path(tmpdir) / "categories.csv"
            csv_file.write_text("category\nA\nA\nB\nC\nA\n")

            result = summarize_tool._run(str(csv_file))

            assert isinstance(result, (str, dict))


class TestDataQualityTools:
    """Integration tests for data quality validation."""

    @pytest.fixture
    def quality_tool(self):
        """Create CheckDataQualityTool instance."""
        from soothe_nano.toolkits.data import CheckDataQualityTool

        return CheckDataQualityTool()

    def test_detect_missing_values(self, quality_tool) -> None:
        """Test detecting missing values in data."""
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_file = Path(tmpdir) / "missing.csv"
            csv_file.write_text("col1,col2\nA,1\nB,\nC,3\n")

            result = quality_tool._run(str(csv_file))

            # Should handle data quality check
            assert isinstance(result, (str, dict))

    def test_detect_duplicates(self, quality_tool) -> None:
        """Test detecting duplicate rows."""
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_file = Path(tmpdir) / "duplicates.csv"
            csv_file.write_text("id,value\n1,A\n2,B\n1,A\n")

            result = quality_tool._run(str(csv_file))

            # Should detect duplicates
            assert isinstance(result, (str, dict))


# ---------------------------------------------------------------------------
# Document Extraction Tools Tests
# ---------------------------------------------------------------------------


class TestDocumentTools:
    """Integration tests for document extraction tools."""

    @pytest.fixture
    def extract_tool(self):
        """Create ExtractTextTool instance."""
        from soothe_nano.toolkits.data import ExtractTextTool

        return ExtractTextTool()

    def test_extract_text_from_txt(self, extract_tool) -> None:
        """Test extracting text from plain text file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            txt_file = Path(tmpdir) / "test.txt"
            txt_file.write_text("This is sample text content.\nLine 2.\nLine 3.")

            result = extract_tool._run(str(txt_file))

            # Should extract text
            assert isinstance(result, str)
            assert "sample text content" in result

    def test_extract_text_from_markdown(self, extract_tool) -> None:
        """Test extracting text from Markdown file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            md_file = Path(tmpdir) / "test.md"
            md_file.write_text("# Header\n\nParagraph text.\n\n- List item")

            result = extract_tool._run(str(md_file))

            assert isinstance(result, str)
            assert "Header" in result


class TestDocumentQA:
    """Integration tests for document Q&A capabilities."""

    @pytest.fixture
    def qa_tool(self):
        """Create AskAboutFileTool instance."""
        from soothe_nano.toolkits.data import AskAboutFileTool

        return AskAboutFileTool()

    @pytest.mark.integration
    def test_ask_about_text_file(self, qa_tool) -> None:
        """Test asking questions about text file content.

        This test makes real LLM API calls and requires API keys.
        Marked as integration to skip in unit test runs.
        """
        import os

        if not (os.getenv("OPENAI_API_KEY") or os.getenv("ANTHROPIC_API_KEY")):
            pytest.skip("LLM API key required for document Q&A")

        with tempfile.TemporaryDirectory() as tmpdir:
            txt_file = Path(tmpdir) / "info.txt"
            txt_file.write_text("Python is a programming language. It was created in 1991.")

            try:
                result = qa_tool._run(str(txt_file), question="When was Python created?")

                # Should answer based on file content
                assert isinstance(result, (str, dict))
            except Exception as e:
                # Skip if model not available or other API issues
                error_str = str(e).lower()
                if "not supported" in error_str or "model" in error_str or "invalid" in error_str:
                    pytest.skip(f"Model or API not available: {e}")
                raise

    def test_ask_about_csv_data(self, qa_tool) -> None:
        """Test asking questions about CSV data."""
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_file = Path(tmpdir) / "sales.csv"
            csv_file.write_text("product,sales\nWidget,1000\nGadget,500\n")

            result = qa_tool._run(str(csv_file), question="What is the total sales?")

            assert isinstance(result, (str, dict))


# ---------------------------------------------------------------------------
# File Info and Metadata Tests
# ---------------------------------------------------------------------------


class TestFileInfoTools:
    """Integration tests for file metadata tools."""

    @pytest.fixture
    def info_tool(self):
        """Create GetDataInfoTool instance."""
        from soothe_nano.toolkits.data import GetDataInfoTool

        return GetDataInfoTool()

    def test_get_csv_file_info(self, info_tool) -> None:
        """Test getting CSV file metadata."""
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_file = Path(tmpdir) / "data.csv"
            csv_file.write_text("a,b,c\n1,2,3\n")

            result = info_tool._run(str(csv_file))

            # Should return file metadata
            assert isinstance(result, (str, dict))
            if isinstance(result, dict):
                assert "size" in result or "rows" in result or "columns" in result

    def test_get_json_file_info(self, info_tool) -> None:
        """Test getting JSON file metadata."""
        with tempfile.TemporaryDirectory() as tmpdir:
            json_file = Path(tmpdir) / "data.json"
            json_file.write_text('{"key": "value"}')

            result = info_tool._run(str(json_file))

            assert isinstance(result, (str, dict))


# ---------------------------------------------------------------------------
# Error Handling Tests
# ---------------------------------------------------------------------------


class TestDataToolErrors:
    """Test error handling for data/document tools."""

    def test_nonexistent_file(self) -> None:
        """Test handling of non-existent file."""
        from soothe_nano.toolkits.data import InspectDataTool

        tool = InspectDataTool()
        result = tool._run("/nonexistent/file.csv")

        # Should return error
        assert isinstance(result, (str, dict))
        if isinstance(result, dict):
            assert "error" in result
        else:
            assert "error" in result.lower()

    def test_unsupported_format(self) -> None:
        """Test handling of unsupported file format."""
        from soothe_nano.toolkits.data import InspectDataTool

        tool = InspectDataTool()

        with tempfile.TemporaryDirectory() as tmpdir:
            unknown_file = Path(tmpdir) / "data.xyz"
            unknown_file.write_text("unknown format")

            result = tool._run(str(unknown_file))

            # Should handle gracefully
            assert isinstance(result, (str, dict))

    def test_corrupted_csv(self) -> None:
        """Test handling of corrupted CSV file."""
        from soothe_nano.toolkits.data import SummarizeDataTool

        tool = SummarizeDataTool()

        with tempfile.TemporaryDirectory() as tmpdir:
            bad_csv = Path(tmpdir) / "bad.csv"
            bad_csv.write_text("not,valid\ncsv,content\ntoo,many,columns\n")

            result = tool._run(str(bad_csv))

            # Should handle parsing errors
            assert isinstance(result, (str, dict))

    def test_empty_file(self) -> None:
        """Test handling of empty file."""
        from soothe_nano.toolkits.data import InspectDataTool

        tool = InspectDataTool()

        with tempfile.TemporaryDirectory() as tmpdir:
            empty_file = Path(tmpdir) / "empty.csv"
            empty_file.write_text("")

            result = tool._run(str(empty_file))

            # Should handle empty file
            assert isinstance(result, (str, dict))
